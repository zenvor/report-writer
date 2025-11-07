import pytest
from openpyxl import Workbook, load_workbook

from weekly_report_writer import (
    _extract_name_from_template,
    _update_weekly_title_text,
    WeeklyReportWriterError,
)


def test_extract_name_with_prefix_style():
    """验证“姓名-第n周周报表”格式可以正确提取姓名"""
    assert _extract_name_from_template("范兴兴-第 n 周周报表") == "范兴兴"


def test_extract_name_with_suffix_style():
    """验证“第n周周报表-姓名”格式可以正确提取姓名"""
    assert _extract_name_from_template("第 n 周周报表-范兴兴") == "范兴兴"


def test_extract_name_with_single_segment():
    """当模板只有一个片段时应直接返回该片段"""
    assert _extract_name_from_template("范兴兴") == "范兴兴"


def test_extract_name_invalid_template():
    """空模板应抛出异常以提醒用户修正命名"""
    with pytest.raises(WeeklyReportWriterError):
        _extract_name_from_template("   ")


def test_update_weekly_title_text_replaces_placeholder(tmp_path):
    """复制模板后应将“第 n 周完成重点工作”替换为具体周次"""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet["A1"] = "第 n 周完成重点工作"
    worksheet["A2"] = "其他字段"
    file_path = tmp_path / "周报.xlsx"
    workbook.save(file_path)
    workbook.close()

    _update_weekly_title_text(str(file_path), 44)

    loaded = load_workbook(file_path)
    sheet = loaded.active
    assert sheet["A1"].value == "第44周完成重点工作"
    assert sheet["A2"].value == "其他字段"
    loaded.close()


def test_update_weekly_title_text_handles_compact_placeholder(tmp_path):
    """无空格的“第n周”格式也需要被替换"""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet["A1"] = "第n周完成重点工作"
    file_path = tmp_path / "周报2.xlsx"
    workbook.save(file_path)
    workbook.close()

    _update_weekly_title_text(str(file_path), 12)

    loaded = load_workbook(file_path)
    sheet = loaded.active
    assert sheet["A1"].value == "第12周完成重点工作"
    loaded.close()
