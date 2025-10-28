import argparse
import os
import shutil
import logging
from datetime import datetime
from pathlib import Path
from typing import Optional, List, Dict, Any

from openai import OpenAI
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.styles import Alignment

from config_manager import config, ConfigurationError
from gitlab_client import GitLabClient, GitLabClientError

logger = logging.getLogger(__name__)

# 常量定义
DEFAULT_WORK_HOURS = 8
DEFAULT_SUMMARY_FALLBACK = "无提交"
MAX_COMMIT_DISPLAY = 10
DEEPSEEK_BASE_URL = "https://api.deepseek.com"  # Deepseek API 基础地址
EXCEL_START_ROW = 3

class ReportUpdaterError(Exception):
    """报告更新器异常"""
    pass

class ExcelOperationError(ReportUpdaterError):
    """Excel操作异常"""
    pass

class AIServiceError(ReportUpdaterError):
    """AI服务异常"""
    pass

class BackupError(ReportUpdaterError):
    """备份操作异常"""
    pass

class ReportUpdater:
    """日报更新器，负责整个日报更新流程"""
    
    def __init__(self):
        self.projects = config.get("gitlab.projects", [])
        self.default_branch = config.get("gitlab.default_branch", "master")
        self.gitlab_client = None  # 在单项目模式下使用

        if not self.projects:
            self.gitlab_client = GitLabClient()

        self.deepseek_api_key = config.get_env_or_config("DEEPSEEK_API_KEY", "deepseek.api_key")
        
        if self.deepseek_api_key:
            self.openai_client = OpenAI(api_key=self.deepseek_api_key, base_url=DEEPSEEK_BASE_URL)
        else:
            self.openai_client = None
        
        # 从配置中获取 Excel 列配置
        self.date_column = config.get("excel_columns.date", 6)
        self.content_column = config.get("excel_columns.content", 7)
        self.hours_column = config.get("excel_columns.hours", 8)
        
        logger.info(f"报告更新器初始化完成 - Excel列配置: 日期={self.date_column}, 内容={self.content_column}, 工时={self.hours_column}")
    
    def update_daily_report(self, xlsx_path: str, date_obj: datetime, work_hours: int = DEFAULT_WORK_HOURS) -> bool:
        """更新日报的主流程"""
        logger.info(f"开始更新日报: {xlsx_path}, 日期: {date_obj.strftime('%Y-%m-%d')}")
        
        try:
            # 1. 验证文件存在
            self._validate_excel_file(xlsx_path)
            
            # 2. 创建备份
            self._create_backup_with_validation(xlsx_path)
            
            # 3. 获取提交信息
            all_commits = self._fetch_all_commits(date_obj)
            
            # 4. 生成日报摘要
            summary = self._generate_summary_with_fallback(all_commits)
            
            # 5. 写入 Excel
            self._write_to_excel_safely(xlsx_path, date_obj, summary, work_hours)
            
            logger.info("日报更新成功")
            return True
            
        except Exception as e:
            logger.error(f"更新日报失败: {e}")
            return False
    
    def _validate_excel_file(self, xlsx_path: str) -> None:
        """验证Excel文件"""
        if not os.path.exists(xlsx_path):
            raise ReportUpdaterError(f"Excel 文件不存在: {xlsx_path}")
        
        if not xlsx_path.lower().endswith('.xlsx'):
            raise ReportUpdaterError(f"文件格式不正确，需要 .xlsx 文件: {xlsx_path}")
        
        # 检查文件是否可读
        try:
            with open(xlsx_path, 'rb') as f:
                f.read(1)
        except PermissionError:
            raise ReportUpdaterError(f"Excel 文件被占用或无权限访问: {xlsx_path}")
    
    def _create_backup_with_validation(self, xlsx_path: str) -> None:
        """创建备份并验证"""
        if not self._create_backup(xlsx_path):
            raise BackupError("创建备份失败")
    
    def _fetch_all_commits(self, date_obj: datetime) -> Dict[str, List[str]]:
        """获取所有项目的提交信息"""
        all_commits = {}

        if self.projects:
            # 多项目模式
            for project in self.projects:
                project_id = project.get("id")
                branch = project.get("branch", self.default_branch)
                client = GitLabClient(project_id=str(project_id), branch=branch)
                
                logger.info(f"正在获取项目 {project_id} (分支: {branch}) 的提交")
                commits = self._fetch_commits_safely(client, date_obj)
                if commits:
                    all_commits[str(project_id)] = commits
        else:
            # 单项目模式
            logger.info("单项目模式，获取提交")
            if self.gitlab_client:
                commits = self._fetch_commits_safely(self.gitlab_client, date_obj)
                if commits:
                    all_commits[self.gitlab_client.project_id] = commits

        return all_commits

    def _fetch_commits_safely(self, client: GitLabClient, date_obj: datetime) -> List[str]:
        """安全地获取单个项目的提交信息"""
        try:
            commits = client.fetch_commits(date_obj)
            logger.info(f"项目 {client.project_id}: 获取到 {len(commits)} 条提交信息")
            return commits
        except GitLabClientError as e:
            logger.warning(f"项目 {client.project_id}: 获取提交信息失败: {e}")
            return []
    
    def _generate_summary_with_fallback(self, all_commits: Dict[str, List[str]]) -> str:
        """生成摘要，带降级处理"""
        if not all_commits:
            return DEFAULT_SUMMARY_FALLBACK

        try:
            summary = self._generate_summary(all_commits)
            if not summary:
                logger.warning("AI生成摘要为空，使用简单摘要")
                summary = self._create_simple_summary_for_all(all_commits)
            
            logger.info(f"生成日报摘要:\n{summary}")
            return summary
        except Exception as e:
            logger.error(f"生成摘要失败: {e}，将使用简单摘要")
            return self._create_simple_summary_for_all(all_commits)

    def _write_to_excel_safely(self, xlsx_path: str, date_obj: datetime, summary: str, work_hours: int) -> None:
        """安全地写入Excel"""
        if not self._write_to_excel(xlsx_path, date_obj, summary, work_hours):
            raise ExcelOperationError("写入Excel失败")
    
    def _create_backup(self, xlsx_path: str) -> bool:
        """创建文件备份"""
        if not config.get("backup.enabled", True):
            logger.debug("备份功能已禁用")
            return True
        
        try:
            backup_dir = Path(xlsx_path).parent / "backups"
            backup_dir.mkdir(exist_ok=True)
            
            # 生成备份文件名
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_filename = f"{Path(xlsx_path).stem}_{timestamp}.xlsx"
            backup_path = backup_dir / backup_filename
            
            # 复制文件
            shutil.copy2(xlsx_path, backup_path)
            logger.info(f"创建备份文件: {backup_path}")
            
            # 清理旧备份
            self._cleanup_old_backups(backup_dir)
            
            return True
            
        except Exception as e:
            logger.error(f"创建备份失败: {e}")
            return False
    
    def _cleanup_old_backups(self, backup_dir: Path) -> None:
        """清理旧的备份文件"""
        try:
            max_backups = config.get("backup.max_backups", 5)
            backup_files = sorted(
                backup_dir.glob("*.xlsx"), 
                key=lambda x: x.stat().st_mtime, 
                reverse=True
            )
            
            for backup_file in backup_files[max_backups:]:
                backup_file.unlink()
                logger.debug(f"删除旧备份文件: {backup_file}")
                
        except Exception as e:
            logger.warning(f"清理旧备份时出错: {e}")
    
    def _generate_summary(self, all_commits: Dict[str, List[str]]) -> str:
        """生成日报摘要"""
        if not all_commits:
            return DEFAULT_SUMMARY_FALLBACK

        if len(all_commits) == 1:
            # 单项目，逻辑不变
            project_id, commits = next(iter(all_commits.items()))
            return self._generate_single_project_summary(commits)

        # 多项目，将所有提交合并后生成统一摘要
        all_commits_merged = []
        for project_id, commits in all_commits.items():
            if commits:
                all_commits_merged.extend(commits)

        if not all_commits_merged:
            return DEFAULT_SUMMARY_FALLBACK

        # 对合并后的提交生成统一摘要
        logger.info(f"为 {len(all_commits)} 个项目的合并提交生成摘要")
        return self._generate_single_project_summary(all_commits_merged)

    def _generate_single_project_summary(self, commits: List[str]) -> str:
        """为单个项目生成摘要"""
        if not self.openai_client:
            logger.warning("未配置 Deepseek API Key，使用简单摘要")
            return self._create_simple_summary(commits)
        
        try:
            logger.info("使用 Deepseek API 生成摘要")
            return self._call_deepseek_api(commits)
        except AIServiceError as e:
            logger.error(f"调用 Deepseek API 失败: {e}")
            return self._create_simple_summary(commits)
    
    def _create_simple_summary_for_all(self, all_commits: Dict[str, List[str]]) -> str:
        """为所有项目创建简单的摘要"""
        if not all_commits:
            return DEFAULT_SUMMARY_FALLBACK

        # 将所有项目的提交合并，避免重复编号
        all_commits_merged = []
        for project_id, commits in all_commits.items():
            if commits:
                all_commits_merged.extend(commits)
        
        if not all_commits_merged:
            return DEFAULT_SUMMARY_FALLBACK
            
        return self._create_simple_summary(all_commits_merged)

    def _create_simple_summary(self, commits: List[str]) -> str:
        """创建简单的摘要"""
        if not commits:
            return "无提交记录"
        
        # 生成条目式摘要
        summary_items = []
        for i, commit in enumerate(commits[:MAX_COMMIT_DISPLAY], 1):
            summary_items.append(f"{i}. {commit}")
        
        if len(commits) > MAX_COMMIT_DISPLAY:
            remaining_count = len(commits) - MAX_COMMIT_DISPLAY
            summary_items.append(f"{MAX_COMMIT_DISPLAY + 1}. 以及其他{remaining_count}项提交")
        
        return "\n".join(summary_items)
    
    def _call_deepseek_api(self, commits: List[str]) -> str:
        """调用 Deepseek API 生成摘要"""
        logger.info("调用 Deepseek API 生成摘要")
        
        prompt = self._create_prompt(commits)
        messages = self._create_api_messages(prompt)
        
        try:
            deepseek_config = config.get("deepseek_config", {})
            response = self.openai_client.chat.completions.create(
                model=deepseek_config.get("model", "deepseek-chat"),
                messages=messages,
                stream=False,
                temperature=deepseek_config.get("temperature", 0.4),
                max_tokens=deepseek_config.get("max_tokens", 300)
            )
            summary = response.choices[0].message.content.strip()
            logger.info("Deepseek API 调用成功")
            return summary
            
        except Exception as e:
            raise AIServiceError(f"API 请求失败: {e}")
    
    def _create_prompt(self, commits: List[str]) -> str:
        """创建API提示词"""
        return (
            "以下是今天的 Git 提交信息，请提炼为精简的日报条目。要求：\n"
            "1. 智能合并相关工作（如：多个修复合并为一条，同一模块的改动合并）\n"
            "2. 提炼为少量核心工作条目（通常3-5条），避免罗列细节\n"
            "3. 格式：1. XXX 2. XXX 3. XXX\n"
            "4. 每条简洁明了，突出关键成果\n\n"
            "Git 提交记录：\n"
            + "\n".join(f"- {commit}" for commit in commits)
        )
    
    def _create_api_messages(self, prompt: str) -> List[Dict[str, str]]:
        """创建API请求消息"""
        deepseek_config = config.get("deepseek_config", {})
        
        return [
            {
                "role": "system", 
                "content": deepseek_config.get(
                    "system_prompt", 
                    "你是一名中国程序员，擅长写精炼的技术日报。请将提交信息总结为最多2句话，每句话不超过30字。"
                )
            },
            {"role": "user", "content": prompt}
        ]
    
    def _write_to_excel(self, xlsx_path: str, date_obj: datetime, summary: str, work_hours: int) -> bool:
        """写入 Excel 文件"""
        try:
            logger.info(f"打开 Excel 文件: {xlsx_path}")
            
            # 修复：openpyxl 不支持上下文管理器
            workbook = load_workbook(xlsx_path)
            worksheet = workbook.active
            
            # 格式化目标日期
            target_date = self._format_target_date(date_obj)
            logger.info(f"查找日期行: {target_date}")
            
            # 查找对应的日期行
            row_found = self._find_and_update_row(worksheet, target_date, summary, work_hours)
            
            if row_found:
                # 保存文件
                workbook.save(xlsx_path)
                logger.info(f"成功写入日期 {target_date} 的日报")
                return True
            else:
                logger.warning(f"未找到日期 {target_date} 对应的行")
                return False
            
        except InvalidFileException as e:
            logger.error(f"Excel 文件格式错误: {e}")
            return False
        except PermissionError as e:
            logger.error(f"Excel 文件被占用或无权限: {e}")
            return False
        except Exception as e:
            logger.error(f"写入 Excel 文件时发生错误: {e}")
            return False
    
    def _format_target_date(self, date_obj: datetime) -> str:
        """格式化目标日期"""
        return date_obj.strftime("%Y/%-m/%-d")
    
    def _find_and_update_row(self, worksheet, target_date: str, summary: str, work_hours: int) -> bool:
        """查找并更新对应的行"""
        for row in range(EXCEL_START_ROW, worksheet.max_row + 1):
            cell_value = worksheet.cell(row, self.date_column).value
            
            if cell_value and str(cell_value).strip() == target_date:
                logger.info(f"找到日期行: 第 {row} 行")
                
                # 写入工作内容并设置自动换行
                content_cell = worksheet.cell(row=row, column=self.content_column)
                content_cell.value = summary
                content_cell.alignment = Alignment(wrap_text=True)
                
                # 如果工时列为空，填入默认工时
                if not worksheet.cell(row, self.hours_column).value:
                    worksheet.cell(row, self.hours_column, work_hours)
                
                return True
        
        return False
    
    def health_check(self) -> Dict[str, bool]:
        """健康检查"""
        logger.info("开始执行健康检查")
        
        status = {
            "gitlab_connection": False,
            "deepseek_api_key": bool(self.deepseek_api_key),
            "config_loaded": True
        }
        
        # 检查 GitLab 连接
        try:
            status["gitlab_connection"] = self.gitlab_client.validate_connection()
        except Exception as e:
            logger.error(f"GitLab 连接检查失败: {e}")
        
        # 检查配置完整性
        try:
            self._validate_configuration()
        except ConfigurationError as e:
            logger.error(f"配置检查失败: {e}")
            status["config_loaded"] = False
        
        logger.info(f"健康检查完成: {status}")
        return status
    
    def _validate_configuration(self) -> None:
        """验证配置完整性"""
        # 验证Excel列配置
        if not all([self.date_column, self.content_column, self.hours_column]):
            raise ConfigurationError("Excel列配置不完整")
        
        # 验证必要的配置项
        required_configs = ["excel_columns", "retry_config", "logging"]
        for config_key in required_configs:
            if not config.get(config_key):
                raise ConfigurationError(f"缺少必要配置: {config_key}")

def main():
    """主函数"""
    parser = argparse.ArgumentParser(description="日报更新器")
    parser.add_argument("--file", required=True, help="月报 Excel 文件路径")
    parser.add_argument("--date", help="日期 YYYY-MM-DD，默认今天")
    parser.add_argument("--hours", type=int, default=DEFAULT_WORK_HOURS, help=f"工作小时数，默认 {DEFAULT_WORK_HOURS}")
    parser.add_argument("--health-check", action="store_true", help="执行健康检查")
    
    args = parser.parse_args()
    
    try:
        # 健康检查
        if args.health_check:
            updater = ReportUpdater()
            status = updater.health_check()
            print("健康检查结果:")
            for key, value in status.items():
                print(f"  {key}: {'✅' if value else '❌'}")
            return
        
        # 解析日期
        date_obj = datetime.strptime(args.date, "%Y-%m-%d") if args.date else datetime.now()
        
        # 创建更新器并执行更新
        updater = ReportUpdater()
        success = updater.update_daily_report(args.file, date_obj, args.hours)
        
        if success:
            print("✅ 日报更新成功")
        else:
            print("❌ 日报更新失败")
            exit(1)
            
    except Exception as e:
        logger.error(f"程序执行失败: {e}")
        print(f"❌ 程序执行失败: {e}")
        exit(1)

if __name__ == "__main__":
    main() 