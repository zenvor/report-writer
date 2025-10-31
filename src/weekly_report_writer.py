"""
周报生成器模块

从月报文件中读取本周日报内容并自动填入周报表格
"""

import logging
import shutil
from datetime import datetime
from pathlib import Path
from typing import List, Optional, Dict
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from date_utils import get_week_start, get_week_dates, is_date_match, format_date_chinese, get_week_range_str

logger = logging.getLogger(__name__)

# 常量定义
MONTHLY_DATE_COLUMN = 6     # 月报中的日期列（F列）
MONTHLY_CONTENT_COLUMN = 7  # 月报中的内容列（G列）
MONTHLY_START_ROW = 3       # 月报数据起始行

WEEKLY_CONTENT_COLUMN = 2   # 周报中的事项列（B列）
WEEKLY_START_ROW = 3        # 周报数据起始行（序号1对应第3行）

# 周报模板相关配置
DEFAULT_TEMPLATE_DIR = "data/weekly report template"  # 周报模板目录
DEFAULT_DATA_DIR = "data"                             # 数据目录


class WeeklyReportWriterError(Exception):
    """周报生成器异常"""
    pass


def find_template_file(template_dir: str = DEFAULT_TEMPLATE_DIR) -> Optional[str]:
    """
    在模板目录中查找周报模板文件

    Args:
        template_dir: 模板目录路径

    Returns:
        模板文件路径，未找到返回None
    """
    template_path = Path(template_dir)

    if not template_path.exists():
        logger.warning(f"模板目录不存在: {template_path}")
        return None

    # 查找.xlsx文件
    template_files = list(template_path.glob("*.xlsx"))

    if template_files:
        # 优先返回包含"周报"的文件
        for file in template_files:
            if "周报" in file.name:
                logger.info(f"找到周报模板: {file}")
                return str(file)

        # 返回第一个找到的文件
        logger.info(f"找到模板文件: {template_files[0]}")
        return str(template_files[0])

    logger.warning("模板目录中未找到.xlsx文件")
    return None


def copy_template_to_data_dir(template_path: str, data_dir: str = DEFAULT_DATA_DIR, week_start_date: Optional[datetime] = None) -> Optional[str]:
    """
    将周报模板从模板目录复制到数据目录

    Args:
        template_path: 模板文件路径
        data_dir: 目标数据目录
        week_start_date: 周一日期，用于生成新文件名

    Returns:
        复制后的文件路径，复制失败返回None
    """
    try:
        template_file = Path(template_path)
        data_path = Path(data_dir)

        if not template_file.exists():
            raise WeeklyReportWriterError(f"模板文件不存在: {template_path}")

        # 确保目标目录存在
        data_path.mkdir(parents=True, exist_ok=True)

        # 生成新文件名
        if week_start_date is None:
            week_start_date = get_week_start()

        # 计算周数（ISO标准），减1
        week_number = week_start_date.isocalendar()[1] - 1

        # 从模板文件名中提取"姓名"部分
        # 模板格式：姓名-第 n 周周报表.xlsx
        template_name = template_file.stem  # 不含扩展名
        name_part = template_name.split('-')[0].strip()  # 提取第一个"-"前的部分作为姓名

        # 新文件名格式：第N周周报表-姓名.xlsx
        # 示例：第44周周报表-范兴兴.xlsx
        new_filename = f"第{week_number}周周报表-{name_part}.xlsx"
        target_path = data_path / new_filename

        # 复制文件
        shutil.copy2(template_file, target_path)
        logger.info(f"复制周报模板成功: {template_file.name} -> {new_filename}")

        return str(target_path)

    except Exception as e:
        logger.error(f"复制周报模板失败: {e}")
        raise WeeklyReportWriterError(f"复制周报模板失败: {e}")


class WeeklyReportWriter:
    """
    周报生成器

    从月报文件中读取本周一到周五的日报内容，
    按时间顺序填入周报Excel的"完成重点工作"表格
    """

    def __init__(self, monthly_report_path: str, weekly_report_path: str, use_template: bool = False, template_dir: str = DEFAULT_TEMPLATE_DIR, week_start_date: Optional[datetime] = None):
        """
        初始化周报生成器

        Args:
            monthly_report_path: 月报文件路径
            weekly_report_path: 周报文件路径（当use_template=False时使用）
            use_template: 是否从模板复制周报（默认False）
            template_dir: 模板目录路径
            week_start_date: 周一日期，用于模板复制时生成新文件名
        """
        self.monthly_report_path = Path(monthly_report_path)
        self.use_template = use_template
        self.template_dir = template_dir
        self.week_start_date = week_start_date

        # 验证月报文件存在
        if not self.monthly_report_path.exists():
            raise WeeklyReportWriterError(f"月报文件不存在: {monthly_report_path}")

        # 处理周报文件路径
        if use_template:
            # 从模板复制周报文件
            template_file = find_template_file(template_dir)
            if not template_file:
                raise WeeklyReportWriterError(f"模板目录中未找到周报模板: {template_dir}")

            # 复制模板到数据目录（weekly_report_path本身就是目录）
            target_data_dir = weekly_report_path if Path(weekly_report_path).is_dir() else str(Path(weekly_report_path).parent)
            copied_path = copy_template_to_data_dir(template_file, data_dir=target_data_dir, week_start_date=week_start_date)
            if not copied_path:
                raise WeeklyReportWriterError("复制周报模板失败")

            self.weekly_report_path = Path(copied_path)
        else:
            # 使用指定的周报文件
            self.weekly_report_path = Path(weekly_report_path)
            if not self.weekly_report_path.exists():
                raise WeeklyReportWriterError(f"周报文件不存在: {weekly_report_path}")

        logger.info(f"周报生成器初始化 - 月报: {self.monthly_report_path.name}, 周报: {self.weekly_report_path.name}")

    def generate_weekly_report(self, week_start_date: Optional[datetime] = None) -> bool:
        """
        生成周报

        Args:
            week_start_date: 周一日期，默认为本周一

        Returns:
            是否成功
        """
        try:
            # 计算本周一
            if week_start_date is None:
                week_start_date = get_week_start()
            else:
                # 确保传入的日期是周一
                week_start_date = get_week_start(week_start_date)

            week_range = get_week_range_str(week_start_date)
            logger.info(f"开始生成周报 - 周期: {week_range}")

            # 读取本周日报内容
            weekly_contents = self._read_weekly_reports(week_start_date)

            # 统计读取结果
            found_count = sum(1 for content in weekly_contents if content is not None)
            logger.info(f"从月报中读取到 {found_count}/5 天的日报内容")

            if found_count == 0:
                logger.warning("未找到任何日报内容，请检查月报文件")
                return False

            # 写入周报
            self._write_to_weekly_report(weekly_contents)

            logger.info("✅ 周报生成成功")
            return True

        except Exception as e:
            logger.error(f"❌ 生成周报失败: {e}", exc_info=True)
            return False

    def _read_weekly_reports(self, monday: datetime) -> List[Optional[str]]:
        """
        从月报中读取本周一到周五的日报内容

        Args:
            monday: 周一日期

        Returns:
            包含5个元素的列表，对应周一到周五的日报内容，未找到的为None
        """
        try:
            workbook = load_workbook(self.monthly_report_path, data_only=True)
            worksheet = workbook.active

            weekly_contents = []
            weekday_names = ["周一", "周二", "周三", "周四", "周五"]

            # 获取本周一到周五的日期
            week_dates = get_week_dates(monday)

            for i, target_date in enumerate(week_dates):
                weekday_name = weekday_names[i]
                date_str = format_date_chinese(target_date)

                # 在月报中查找对应日期
                content = None
                for row_idx in range(MONTHLY_START_ROW, worksheet.max_row + 1):
                    cell_date = worksheet.cell(row=row_idx, column=MONTHLY_DATE_COLUMN).value

                    if is_date_match(cell_date, target_date):
                        content = worksheet.cell(row=row_idx, column=MONTHLY_CONTENT_COLUMN).value

                        if content:
                            content = str(content).strip()  # 清理空白字符
                            logger.info(f"  ✓ {date_str}: 找到日报内容 (月报第{row_idx}行)")
                        else:
                            logger.warning(f"  ⚠ {date_str}: 日报内容为空 (月报第{row_idx}行)")
                            content = None
                        break

                if content is None:
                    logger.warning(f"  ✗ {date_str}: 未找到日报")

                weekly_contents.append(content)

            workbook.close()
            return weekly_contents

        except InvalidFileException as e:
            raise WeeklyReportWriterError(f"无法打开月报文件，可能格式不正确: {e}")
        except Exception as e:
            raise WeeklyReportWriterError(f"读取月报文件失败: {e}")

    def _write_to_weekly_report(self, contents: List[Optional[str]]) -> None:
        """
        将日报内容写入周报

        Args:
            contents: 包含5个元素的列表，对应周一到周五的日报内容
        """
        try:
            workbook = load_workbook(self.weekly_report_path)
            worksheet = workbook.active

            weekday_names = ["周一", "周二", "周三", "周四", "周五"]
            write_count = 0

            # 填写内容到对应序号（保持序号1-5对应周一到周五）
            for i, content in enumerate(contents):
                sequence_number = i + 1  # 序号1-5
                target_row = WEEKLY_START_ROW + i  # B3-B7行

                if content:
                    # 写入内容
                    cell = worksheet.cell(row=target_row, column=WEEKLY_CONTENT_COLUMN)
                    cell.value = content

                    # 设置自动换行（与日报保持一致）
                    from openpyxl.styles import Alignment
                    cell.alignment = Alignment(wrap_text=True, vertical='top')

                    logger.info(f"  ✓ 序号{sequence_number}({weekday_names[i]}): 写入周报第{target_row}行")
                    write_count += 1
                else:
                    # 保持该行为空（不覆盖已有内容）
                    logger.debug(f"  - 序号{sequence_number}({weekday_names[i]}): 无内容，跳过")

            # 保存文件
            workbook.save(self.weekly_report_path)
            workbook.close()

            logger.info(f"成功写入 {write_count}/5 天的日报到周报文件")

        except PermissionError:
            raise WeeklyReportWriterError(f"无法写入周报文件，文件可能被占用: {self.weekly_report_path}")
        except Exception as e:
            raise WeeklyReportWriterError(f"写入周报文件失败: {e}")

    def preview_weekly_report(self, week_start_date: Optional[datetime] = None) -> Dict[str, Optional[str]]:
        """
        预览周报内容（不写入文件）

        Args:
            week_start_date: 周一日期，默认为本周一

        Returns:
            字典，键为"周一"到"周五"，值为对应的日报内容
        """
        if week_start_date is None:
            week_start_date = get_week_start()
        else:
            week_start_date = get_week_start(week_start_date)

        weekly_contents = self._read_weekly_reports(week_start_date)
        weekday_names = ["周一", "周二", "周三", "周四", "周五"]

        return {
            weekday_names[i]: content
            for i, content in enumerate(weekly_contents)
        }
